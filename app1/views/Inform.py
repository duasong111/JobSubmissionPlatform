from django.shortcuts import render, HttpResponse, redirect
from django.views.decorators.csrf import csrf_exempt
from app1.utils.bootstrap import BootStrapModelForm
from django.http import JsonResponse
from app1 import models
from openpyxl import load_workbook
from datetime import datetime
import random
class InformModelForm(BootStrapModelForm):
    class Meta:
        model = models.Work_inform
        fields = "__all__"
        exclude = ["create_time"]  # 将这个内容不显示
def Inform_list(request):
    # 按照时间顺序来进行最新数据输出
    queryset = models.Work_inform.objects.all().order_by("-id")

    form = InformModelForm()
    context = {
        "form": form,
        "queryset": queryset
    }
    return render(request, 'Inform_Homework.html', context)
# 上边俩是将数据显示在数据库中
@csrf_exempt
def Inform_add(request):
    """新建订单"""
    form = InformModelForm(data=request.POST)
    if form.is_valid():
        # 下面的那个是自动生成的数然后填充进去的数值
        # 随机函数获取今日时间，然后添加到数据中
        form.instance.create_time = datetime.now().strftime("%Y-%m-%d")
        # 去找到登录的这个账户的管理员，然后来进行控制
        # form.instance.admin_id = request.session["info"]["id"]
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": True, "error": form.errors})
def Inform_delete(request):
    """删除订单"""
    uid = request.GET.get("uid")
    exists = models.Work_inform.objects.filter(id=uid).exists()
    if not exists:
        return JsonResponse({"status": False, 'error': "获取失败,数据不存在"})
    models.Work_inform.objects.filter(id=uid).delete()
    return JsonResponse({"status": True})
def Inform_detail(request):
    uid = request.GET.get("uid")
    row_dict = models.Work_inform.objects.filter(id=uid).values("inform").first()
    if not row_dict:
        return JsonResponse({"status": False, 'error': "数据不存在"})
    result = {
        "status": True,
        "data": row_dict
    }
    return JsonResponse(result)
# 对编辑页面的重新规划
@csrf_exempt
def Inform_edit(request):
    """编辑页面"""
    uid = request.GET.get("uid")
    row_object = models.Work_inform.objects.filter(id=uid).first()
    if not row_object:
        return JsonResponse({"status": False, 'tips': "数据不存在"})
    form = InformModelForm(data=request.POST, instance=row_object)
    if form.is_valid():
        form.save()
        return JsonResponse({"status": True})
    return JsonResponse({"status": True, "error": form.errors})
#饼状图的展示
def Inform_pie(request):
    """饼状图"""
    db_data_list = [
        {"value": 2048, "name": '已交'},
        {"value": 735, "name": '未交'},
        {"value": 1580, "name": '展示'},
    ]

    result = {
        "status": True,
        "data": db_data_list,
    }
    return JsonResponse(result)

#采用表格的形式去进行作业的提交
def Inform_multi(request):
    """批量上传excel文件"""
    # 1首先获取文件的对象
    file_objects = request.FILES.get("exc")
    # print(file_objects.name) #获取文件的名字

    wb = load_workbook(file_objects)
    sheet = wb.worksheets[0]
    dayTime = datetime.now().strftime("%Y-%m-%d")
    for row in sheet.iter_rows(min_row=2):
        text = row[0].value
        text2= row[1].value
        print("{}--{}".format(text,text2))
        exist = models.Work_inform.objects.filter(inform=text).exists()
        if not exist:
            models.Work_inform.objects.create(inform=text, level=text2, create_time=datetime.now())
            print("{}-{}-{}".format(text,text2,datetime.now()))
        else:
            return  HttpResponse("重复上传或格式不正确，请重新提交")
    return redirect("/Inform/Homework/")

